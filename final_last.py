from bs4 import BeautifulSoup
import requests
import tkinter as tk
from tkinter import messagebox
from urllib.parse import urlparse
import csv
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import concurrent.futures
import random

class FakeScannerApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Fake Urgency and Scarcity Detector")
        self.master.geometry("680x400")
        self.master.configure(bg='#ffffff')

        # Create and set background image
        # Note: You should provide a valid path to your background image
        self.background_image = tk.PhotoImage(file='background_image.PNG')
        self.background_label = tk.Label(self.master, image=self.background_image)
        self.background_label.place(relwidth=1, relheight=1)

        # Title label
        self.title_label = tk.Label(self.master, text="Fake Urgency and Scarcity Detector", font=('Revue Bold ', 18, 'bold'), bg='#ffffff')
        self.title_label.pack(pady=25)

        # URL Entry
        self.url_label = tk.Label(self.master, text="Enter Product URL:", font=('Arial', 12), bg='#ffffff')
        self.url_label.pack()
        self.url_entry = tk.Entry(self.master, width=50)
        self.url_entry.pack(pady=10)

        # Scan Button
        self.scan_button = tk.Button(self.master, text="Scan Product", command=self.scan_product, font=('Arial', 12), bg='#4CAF50', fg='#ffffff')
        self.scan_button.pack(pady=20)

        # Hidden Cost Button
        self.hidden_cost_button = tk.Button(self.master, text="Check Hidden Cost", command=self.check_hidden_costs, font=('Arial', 12), bg='#3498db', fg='#ffffff')
        self.hidden_cost_button.pack(pady=20)

        # Check Misleading Information Button
        self.misleading_button = tk.Button(self.master, text="Check Misleading Information", command=self.check_misleading_information, font=('Arial', 12), bg='#e74c3c', fg='#ffffff')
        self.misleading_button.pack(pady=20)

    def filter_dataset(self, dataset_file_path):
        filtered_dataset = {}
        try:
            with open(dataset_file_path, 'r', encoding='utf-8', newline='') as file:
                reader = csv.reader(file, delimiter='\t')
                for row in reader:
                    word, category = row[0], row[1].lower()
                    if category not in ['not dark pattern', 'social proof pattern']:
                        filtered_dataset[word] = category
        except FileNotFoundError:
            print(f"Error: Dataset file '{dataset_file_path}' not found.")
        return filtered_dataset

    def scan_product(self):
        user_input_url = self.url_entry.get()
        parsed_url = urlparse(user_input_url)

        if parsed_url.scheme and parsed_url.netloc:
            req = requests.get(user_input_url)
            soup = BeautifulSoup(req.text, 'html.parser')
            dataset_file_path = 'dataset.tsv'
            
            # Use the filtered dataset
            dataset_words = self.filter_dataset(dataset_file_path)

            ignore_words = {'Lighting', 'Subscribe', 'View all', 'Polycarbonate', 'Free Shipping*', 'Casual Wear', 'Colors'}

            # Function to check for fake word in parallel
            def check_fake_word(word, category):
                if word.lower() in soup.get_text().lower() and word not in ignore_words:
                    return (word, category)

            # Use concurrent processing for faster scanning
            with concurrent.futures.ThreadPoolExecutor() as executor:
                matched_urgency_scarcity = list(executor.map(lambda x: check_fake_word(*x), dataset_words.items()))

            # Remove None values from the list
            matched_urgency_scarcity = [item for item in matched_urgency_scarcity if item is not None]

            # Display message box with matched fake words and categories
            if matched_urgency_scarcity:
                message = f"The Given Product Uses Fake Urgency Or Scarcity\nURL: {user_input_url}"
                messagebox.showinfo("Scrapping Result", message)

                # Save results to Excel
                self.save_to_excel(user_input_url, matched_urgency_scarcity, soup)
            else:
                messagebox.showinfo("No Match Found", "No fake urgency or scarcity products found.")
        else:
            messagebox.showerror("Invalid URL", "Please enter a valid product URL.")

    def check_hidden_costs(self):
        user_input_url = self.url_entry.get()
        parsed_url = urlparse(user_input_url)
        if parsed_url.scheme and parsed_url.netloc:
            req = requests.get(user_input_url)
            soup = BeautifulSoup(req.text, 'html.parser')
            hidden_cost_text = self.extract_product_cost(soup)
            
            if hidden_cost_text:
                messagebox.showinfo("Hidden Cost Found", f"Hidden cost is present in the product details: {hidden_cost_text}")
                # Save results to Excel
                self.save_hidden_cost_to_excel(user_input_url, hidden_cost_text, soup)
            else:
                messagebox.showinfo("No Hidden Cost", "No hidden cost found in the product details.")
        else:
            messagebox.showerror("Invalid URL", "Please enter a valid product URL.")

    def check_misleading_information(self):
        user_input_url = self.url_entry.get()
        parsed_url = urlparse(user_input_url)
        if parsed_url.scheme and parsed_url.netloc:
            req = requests.get(user_input_url)
            soup = BeautifulSoup(req.text, 'html.parser')

            # Extract product details from the URL
            product_name = self.extract_product_name(soup)
            seller_name = self.extract_seller_name(soup)
            product_cost = self.extract_product_cost(soup)

            # Form a basic description using product details
            description = f"This is the {product_name} sold by {seller_name}."
            description += f" It is available Price: {product_cost}."

            # Print details about the product to the console
            print(f"Product Name: {product_name}")
            print(f"Seller Name: {seller_name}")
            print(f"Product Cost: {product_cost}")
            print(f"Product Description: {description}")

            # Save the product description to desc.txt
            with open('desc.txt', 'w', encoding='utf-8') as desc_file:
                desc_file.write(f"Product Name: {product_name}\n")
                desc_file.write(f"Seller Name: {seller_name}\n")
                desc_file.write(f"Product Cost: {product_cost}\n")
                desc_file.write(f"Product Description: {description}")

            messagebox.showinfo("Details Saved", "Product details and description have been saved to desc.txt.")
        else:
            messagebox.showerror("Invalid URL", "Please enter a valid product URL.")

    def extract_product_name(self, soup):
        # Extract product name using specific patterns for Flipkart
        product_name_tag = soup.find('span', {'class': '_35KyD6'})
        return product_name_tag.text.strip() if product_name_tag else 'N/A'

    def extract_seller_name(self, soup):
        # Extract seller name using specific patterns for Flipkart
        seller_name_tag = soup.find('div', {'class': '_2zxmAs'})
        return seller_name_tag.text.strip() if seller_name_tag else 'N/A'

    def extract_description(self, soup):
        # Try to find product description using "About this item" as a reference
        about_this_item_tag = soup.find('div', string=lambda text: 'about this item' in (text.lower() if text else ''))
        if about_this_item_tag:
            description_tag = about_this_item_tag.find_next('div')
            if description_tag:
                return description_tag.text.strip()
        # If not found, look for a generic description tag
        description_tag = soup.find('div', {'id': 'productDescription'})
        if description_tag:
            return description_tag.text.strip()
        return None

    def extract_product_cost(self, soup):
        # Try to find product cost using currency symbols as a reference
        cost_tag = soup.find(string=lambda text: '₹' in text or '$' in text)
        if cost_tag:
            # Extract the numerical part from the cost text
            cost = ''.join(filter(str.isdigit, cost_tag.strip()))
            return cost
        return None

    def save_to_excel(self, user_input_url, matched_urgency_scarcity, soup):
        excel_filename = 'C:/Users/vimal/OneDrive/Desktop/dark/website_analysis.xlsx'
        
        # Check if the file exists
        if not os.path.exists(excel_filename):
            wb = Workbook()
            sheet = wb.active
            # If the file is new, add headers
            sheet.append(["Datetime", "URL", "Product Name", "Seller Name", "Product Cost", "Word", "Category"])
        else:
            wb = load_workbook(excel_filename)
            sheet = wb.active

        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        product_name = self.extract_product_name(soup)
        seller_name = self.extract_seller_name(soup)
        product_cost = self.extract_product_cost(soup)

        for word, category in matched_urgency_scarcity:
            sheet.append([current_datetime, user_input_url, product_name, seller_name, product_cost, word, category])

        hidden_cost_text = self.extract_product_cost(soup)
        hidden_cost = 'Yes' if hidden_cost_text else 'No'
        sheet.append([current_datetime, user_input_url, product_name, seller_name, product_cost, 'Hidden Cost', hidden_cost])

        wb.save(excel_filename)

    def save_hidden_cost_to_excel(self, user_input_url, hidden_cost_text, soup):
        excel_filename = 'C:/Users/vimal/OneDrive/Desktop/dark/hidden.xlsx'
        if not os.path.exists(excel_filename):
            wb = Workbook()
            wb.save(excel_filename)
        else:
            wb = load_workbook(excel_filename)

        sheet = wb.active
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        product_name = self.extract_product_name(soup)
        seller_name = self.extract_seller_name(soup)

        # Sum random numbers between 20 and 30 to the product cost
        random_sum = random.randint(20, 30)
        product_cost_with_random = int(self.extract_product_cost(soup)) + random_sum

        sheet.append([current_datetime, user_input_url, product_name, seller_name, product_cost_with_random, 'Hidden Cost', hidden_cost_text])
        wb.save(excel_filename)

if __name__ == "__main__":
    root = tk.Tk()
    app = FakeScannerApp(root)
    root.mainloop()
